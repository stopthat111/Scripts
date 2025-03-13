import openpyxl
import csv

def excel_to_csv(input_file, output_file):
    # Load the workbook
    workbook = openpyxl.load_workbook(input_file, data_only=True)

    # Get the sheet names and remove the first sheet
    sheets = workbook.sheetnames
    first_sheet_name = sheets[0]
    del workbook[first_sheet_name]
    print(f"Deleted the first sheet: {first_sheet_name}")  # Debugging output

    # Open the output CSV file for writing
    with open(output_file, mode='w', newline='', encoding='utf-8') as csv_file:
        csv_writer = csv.writer(csv_file)

        # Iterate through the remaining sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Write rows to the CSV
            for row in sheet.iter_rows(values_only=True):
                csv_writer.writerow(row)

    print(f"Excel file '{input_file}' has been converted to CSV '{output_file}'.")

# Example usage
input_excel_file = r"C:\temp\TESTING\NAME.xlsx"
output_csv_file = r"C:\temp\TESTING\NAME.csv"

excel_to_csv(input_excel_file, output_csv_file)
